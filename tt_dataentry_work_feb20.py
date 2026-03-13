# Timetable project
# Intermediate Stage
from tkinter import*
from tkinter import ttk
import tkinter as tk
from openpyxl import load_workbook,Workbook
import pandas as pd
import numpy as np
from tt_show_class_jan20 import TimeTable
import os
def create_excelsheet_orread(file_name,sheet_name):
    if os.path.exists(file_name):
        sheet_names_list = pd.ExcelFile(file_name).sheet_names
        if sheet_name not in sheet_names_list:
            df_init = pd.DataFrame(data_ref)
            with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                df_init.to_excel(writer,sheet_name=sheet_name,index=False,engine = "openpyxl")
            df = pd.read_excel(file_name,sheet_name=sheet_name,engine="openpyxl")
            return(df)
        else: 
            df = pd.read_excel(file_name,sheet_name=sheet_name,engine="openpyxl")
            return(df)
    else:
        df_init = pd.DataFrame(data_ref)
        df_init.to_excel(file_name,sheet_name,index=False,engine = "openpyxl")
        df = pd.read_excel(file_name,sheet_name=sheet_name,engine="openpyxl")
        return(df)   
def find_key(dictionary,value):
    for key,val in dictionary.items():
        if val == value:
            return(key)
    return None
# Combining sublist in data frame and eliminating common ref elements from two lists
def common_lst(df_ref,lstname_ref,lstname_sub):  
    comn_ref_lst = []
    comn_sub_lst = []
    for i in range(len(df_ref)):
        if len(df_ref.at[df_ref.index.tolist()[i],lstname_ref].split(",")) > 1:
            ref_lst = df_ref.at[df_ref.index.tolist()[i],lstname_ref].split(",")
            sub_lst = df_ref.at[df_ref.index.tolist()[i],lstname_sub].split(",")
        else:
            ref_lst = [str(df_ref.at[df_ref.index.tolist()[i],lstname_ref])]
            sub_lst = [str(df_ref.at[df_ref.index.tolist()[i],lstname_sub])]
        comn_ref_lst += ref_lst
        comn_sub_lst += sub_lst                   
    unifrom_ref_list = []
    unifrom_sub_list = []
    for i in range(len(comn_ref_lst)):
        if comn_ref_lst[i] not in unifrom_ref_list:
            unifrom_ref_list.append(comn_ref_lst[i])
            unifrom_sub_list.append(comn_sub_lst[i])
    #print(unifrom_ref_list)
    #print(unifrom_sub_list)                   
    return(unifrom_sub_list)
def excel_read1(file_name,sheet_name):
    data_dfrow = []
    lst_dfcol = []
    workbook = load_workbook(filename=f"{file_name}.xlsx")
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rowdf_dict = dict(zip(headers, row))
        data_dfrow.append(rowdf_dict)
    for column in sheet.iter_cols(min_row=2, values_only=True):
        df_col = []
        for cell in column:
            df_col.append(cell)
        lst_dfcol.append(df_col)
        coldf_dict = dict(zip(headers, lst_dfcol))
    workbook.close()
    row_df_ret = [data_dfrow,coldf_dict,headers]
    return row_df_ret
def excel_writerows(file_name,sheet_name,data_header,data_column):
    workbook = load_workbook(filename=f"{file_name}.xlsx")
    sheet = workbook[sheet_name]
    print(sheet.title)
    return
def parent_dept_selection(event):
    global df_dept
    action_code.set('')
    action_code.config(values=(''))
    dept_emp_code.set('')
    #df_sel = create_excelsheet_orread(file_name='output_data.xlsx',sheet_name=dept_code.get())
    df_dept = create_excelsheet_orread(file_name='faculty_data.xlsx',sheet_name=dept_code.get())
    dept_emp_code.config(values=df_dept.emp_code.tolist())
def emp_code_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    ltp_code.set('')
    ltp_code.config(values=(''))
    assign_dept_code.set('')
    assign_dept_code.config(values=("CE","ME","EE"))
def dept_option_selection(event):
    odd_even_code.config(values=("ODD","EVEN"))
    action_code.set('')
    action_code.config(values=('')) 
def odd_even_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    odd_even = odd_even_code.get()
    if odd_even == 'ODD':
        odd_even_list = ['s1','s3','s5','s7']
    elif odd_even == 'EVEN':
        odd_even_list = ['s2','s4','s6','s8']
    else:
        pass
    semester_code.set('')
    semester_code.config(values=odd_even_list)
def semester_option_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    ltp_code.set('')
    ltp_code.config(values=("Lecture","Tutorial","Practical","Projects"))
def ltp_option_selection(event):
    global df_selected_data,faculty_name
    action_code.set('')
    action_code.config(values=(''))
    global data_faculty,data_tt,data_temp,faculty_sel
    data_curi = excel_read1(f"curiculam_{assign_dept_code.get()}",semester_code.get())
    new_row_lst = [data_curi[0][i] for i in range(len(data_curi[1]["Select"])) if data_curi[1]["Select"][i] != 0]
    data_col_lst = list(data_curi[1].values())  
    cur_col_lst = [[data_col_lst[j][i] for i in range(len(data_curi[1]["Select"])) if data_curi[1]["Select"][i] != 0] for j in range(len(data_curi[2]))]
    cur_col_dict = dict(zip(data_curi[2], cur_col_lst))
    data_tt = excel_read1(f"timeTable_{assign_dept_code.get()}",semester_code.get())
    #data_temp = excel_read1(f"faculty_assignment_{odd_even_code.get()}",assign_dept_code.get())
    #faculty_sel = (data_faculty[1]["nick_name"][(data_faculty[1]["emp_code"].index(int(dept_emp_code.get())))]) #from faculty file
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    global df_dept,df_curi_lec,df_curi_tut
    faculty_name = df_dept.nick_name.tolist()[df_dept.emp_code.tolist().index(int(dept_emp_code.get()))]
    df_curi = create_excelsheet_orread(f"curiculam_{assign_dept_code.get()}.xlsx",semester_code.get())
    df_curi_active = df_curi[df_curi["Select"] != 0]
    df_sel = create_excelsheet_orread(f"faculty_assignment_{odd_even_code.get()}.xlsx",sheet_name=assign_dept_code.get())
    match ltp_code.get():
        case "Lecture":
            df_curi_lec = df_curi_active[df_curi_active["L"] != 0] 
            df_sel_lec = df_sel[["nick_name","deptL_class","semL","L_code"]].copy()         
            print(df_sel_lec)
            if df_sel_lec.empty:
                code_lst = df_curi_lec.Code.tolist()
            else:
                df_sel_lec = df_sel[["nick_name","deptL_class","semL","L_code"]].copy()
                df_sel_lec.dropna(subset=['nick_name'],inplace=True) # remove rows with empty nick_name
                if faculty_name in df_sel_lec.nick_name.tolist():
                    df_sel_lec=df_sel_lec[df_sel_lec["nick_name"]!=faculty_name]
                    code_lst_temp=df_curi_lec.Code.tolist()
                    for i in range(len(common_lst(df_sel_lec,"L_code","L_code"))):    
                        if assign_dept_code.get() == common_lst(df_sel_lec,"L_code","deptL_class")[i] and semester_code.get() == common_lst(df_sel_lec,"L_code","semL")[i]:
                            code_lst_temp.remove(common_lst(df_sel_lec,"L_code","L_code")[i])
                else:
                    code_lst_temp=df_curi_lec.Code.tolist()
                    for i in range(len(common_lst(df_sel_lec,"L_code","L_code"))): 
                        if assign_dept_code.get() == common_lst(df_sel_lec,"L_code","deptL_class")[i] and semester_code.get() == common_lst(df_sel_lec,"L_code","semL")[i]:
                            code_lst_temp.remove(common_lst(df_sel_lec,"L_code","L_code")[i])
                df_sel_mylec = df_sel[["nick_name","L_code","L_class","L_hr"]].copy()
                df_sel_mylec.dropna(subset=['nick_name'],inplace=True) # remove rows with empty nick_name
                if faculty_name in df_sel_mylec.nick_name.tolist():
                    df_sel_mylec=df_sel_mylec[df_sel_mylec["nick_name"]==faculty_name]
                    for i in range(len(common_lst(df_sel_mylec,"L_code","L_code"))):
                        if common_lst(df_sel_mylec,"L_code","L_class")[i] == common_lst(df_sel_mylec,"L_code","L_hr")[i]:
                            code_lst_temp.remove(common_lst(df_sel_mylec,"L_code","L_code")[i])
                code_lst=code_lst_temp    
                print(code_lst)
        case "Tutorial":
            df_curi_tut = df_curi_active[df_curi_active["T"] != 0]
            df_sel_tut = df_sel[["nick_name","deptT_class","semT","T_code"]].copy()
            print(df_sel_tut)
            if df_sel_tut.empty:
                code_lst = df_curi_tut.Code.tolist()
                print(df_curi_tut)
            else:
                df_sel_tut = df_sel[["nick_name","deptT_class","semT","T_code"]].copy()
                df_sel_tut.dropna(subset=['nick_name'],inplace=True) # remove rows with empty nick_name
                print("do the job")
                T_count = []
                print(df_curi_tut.Code.tolist())
                for i in range(len(df_curi_tut)):
                    if list(df_curi_tut.Number)[i] <= 26:
                        T_need = 1
                    elif list(df_curi_tut.Number)[i] <= 48:
                        T_need = 2
                    else:
                        T_need = 3
                    T_count.append(T_need)
                print(T_count)




            return
            faculty_in_tut_sel = df_selected_data[df_selected_data['nick_name'] == faculty_name]
            if list(faculty_in_tut_sel.T_code) != []:
                df_curi_tut_l0 = df_curi_tut[df_curi_tut['Code'] != list(faculty_in_tut_sel.T_code)[0]]
                df_curi_tut = df_curi_tut_l0
            print(list(df_curi_tut.Code))
            T_ref = {}
            for i in range(len(df_curi_tut)):
                if list(df_curi_tut.Number)[i] <= 26:
                    T_need = 1
                elif list(df_curi_tut.Number)[i] <= 48:
                    T_need = 2
                else:
                    T_need = 3
                T_ref[list(df_curi_tut.Code)[i]] = T_need
            print(T_ref)
            print(df_selected_data)
            for i in range(len(df_selected_data)):
                selected_data_curi_key = find_key(df_curi_tut.Code,list(df_selected_data.T_code)[i])
                if selected_data_curi_key == None:
                    continue
                else:
                    if T_ref[list(df_selected_data.T_code)[i]] - list(df_selected_data.T_count)[i] == 0:
                        df_curi_tut_l0 = df_curi_tut[df_curi_tut['Code'] != df_curi_tut.Code[selected_data_curi_key]]  
                        df_curi_tut = df_curi_tut_l0
            code_lst = list(df_curi_tut.Code)
            print(code_lst) 
        case "Practical":
            df_curi_lab = df_curi_active[df_curi_active["P"] != 0]
            faculty_in_lab_sel = df_selected_data[df_selected_data['nick_name'] == faculty_name]
            if list(faculty_in_lab_sel.P_code) != []:
                df_curi_lab_l0 = df_curi_lab[df_curi_lab['Code'] != list(faculty_in_lab_sel.P_code)[0]]
                df_curi_lab = df_curi_lab_l0
            print(list(df_curi_lab.Code))          
            P_ref = {}
            for i in range(len(df_curi_lab)):
                if list(df_curi_lab.Number)[i] <= 18:
                    P_need = 1
                elif list(df_curi_lab.Number)[i] <= 32:
                    P_need = 2
                elif list(df_curi_lab.Number)[i] <= 48:
                    P_need = 3
                else:
                    P_need = 4
                P_ref[list(df_curi_lab.Code)[i]] = P_need         
            print(P_ref)
            print(df_selected_data)
            for i in range(len(df_selected_data)):
                selected_data_curi_key = find_key(df_curi_lab.Code,list(df_selected_data.P_code)[i])
                if selected_data_curi_key == None:
                    continue
                else:
                    if P_ref[list(df_selected_data.P_code)[i]] - list(df_selected_data.P_count)[i] == 0:
                        df_curi_lab_l0 = df_curi_lab[df_curi_lab['Code'] != df_curi_lab.Code[selected_data_curi_key]]  
                        df_curi_lab = df_curi_tut_l0
            code_lst = list(df_curi_lab.Code)
        case "Projects":
            code_lst = ["Guide-4","Guide-8","Other HoD","other CT","Special"]
            faculty_in_guide_sel = df_selected_data[df_selected_data['nick_name'] == faculty_name]
            if list(faculty_in_guide_sel.R_code) != []:
                code_lst_l0 = code_lst
                code_lst = code_lst_l0
            print(code_lst)
            for i in range(len(code_lst)):
                R_ref = [round(int(list(df_curi_active.Number)[i])/4),round(int(list(df_curi_active.Number)[i])/8),1,1,2]
            print(R_ref)
        case _:
            print("Who")
    sub_code.set('')
    sub_code.config(values=code_lst)
def sub_option_selection(event):
    #data_tt["sub_code"] = sub_code.get()
    action_code.config(values=("PROCEED","CLEAR"))
def action_option_selection(event):
    global data_tt,data_faculty
    global class_x,data,ro_co_dict,row_val
    action_entry=action_code.get()
    #print(data_tt)
    if action_entry == "CLEAR":
        lf1_lf2_clear()
        pass
    elif action_entry == "PROCEED":
        TimeTable(root,f"timeTable_{assign_dept_code.get()}.xlsx",semester_code.get())
        ro_co_dict = {}
        for i in range(5):
            ro_co_dict.update({i:[0 for _ in range(7)]})
        week_click(None)
        refresh_button.config(bg="Blue",fg="white")
        text_box2.delete('1.0','end')
        text_box2.insert('1.0',f"Select weekday and one from available slot before REFRESH")
def week_click(week_value:any):
    cb1.config(state='disabled')
    cb2.config(state='disabled')
    cb3.config(state='disabled')
    cb4.config(state='disabled')
    cb5.config(state='disabled')
    cb6.config(state='disabled')
    cb7.config(state='disabled')
    sr0.set(0)
    sr1.set(0)
    sr2.set(0)
    sr3.set(0)
    sr4.set(0)
    sr5.set(0)
    sr6.set(0)
    if data_tt[0][r3.get()][data_tt[2][1]] == "FREE":
        cb1.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][2]] == "FREE":
        cb2.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][3]] == "FREE":
        cb3.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][4]] == "FREE":
        cb4.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][5]] == "FREE":
        cb5.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][6]] == "FREE":
        cb6.config(state='normal')
    verify_button.config(bg="white",fg="black")
def refresh_weekslot():
    ro_co_dict[r3.get()] = [sr0.get(),sr1.get(),sr2.get(),sr3.get(),sr4.get(),sr5.get(),sr6.get()]
    if 1 not in ro_co_dict[r3.get()]:
        text_box2.insert('1.0',f"!!WARNING!! Select weekday and one from available slot before REFRESH...................")
        return
    #print(r3.get(),ro_co_dict[r3.get()].index(1))
    refresh_button.config(bg="yellow",fg="black")
    verify_button.config(bg="Blue",fg="white")
    update_button.config(bg="white",fg="black")
    text_box2.delete('1.0','end')
    text_box2.insert('1.0',f"Confirm your choice {ltp_code.get()} : for the subject {sub_code.get()} for {semester_code.get()} in {assign_dept_code.get()} .... ")
    text_box2.insert('2.0',f"Selected choice {data_tt[1]["week_day"][r3.get()]} for the slot {data_tt[2][(ro_co_dict[r3.get()].index(1))+1]} ")
def verify_ok():
    update_button.config(bg="Blue",fg="white")
    verify_button.config(bg="white",fg="black")
    refresh_button.config(bg="white",fg="black")
    text_box2.delete('1.0','end')
    text_box2.insert('1.0',f"Confirm and SAVE to Time Table")
def update_tt():
    global df_selected_data,faculty_name,df_curi_lec
    data_tt = excel_read1(f"timeTable_{assign_dept_code.get()}",semester_code.get())
    data_faculty = excel_read1("faculty_data",dept_code.get())
    data_temp = excel_read1(f"faculty_assignment_{odd_even_code.get()}",assign_dept_code.get())
    faculty_ref = data_faculty[1]["nick_name"][data_faculty[1]["emp_code"].index(int(dept_emp_code.get()))]
    update_button.config(bg="white",fg="black")
    df_dept_data = pd.read_excel("faculty_data.xlsx",sheet_name=dept_code.get(), usecols=["emp_code","nick_name"], engine = "openpyxl")
    df_data_tt = pd.read_excel(f"timeTable_{assign_dept_code.get()}.xlsx",semester_code.get(),engine = "openpyxl")
    df_sel = create_excelsheet_orread(f"faculty_assignment_{odd_even_code.get()}.xlsx",sheet_name=assign_dept_code.get())
    match ltp_code.get():
        case "Lecture":
            class_hr_index = df_curi_lec.Code.tolist().index(sub_code.get())
            lec_class_hr = str(df_curi_lec.L.tolist()[class_hr_index])
            print(class_hr_index,lec_class_hr)
            if faculty_name in list(df_sel.nick_name):
                index_new = list(df_sel.nick_name).index(faculty_name)
                if sub_code.get() in df_sel.at[index_new,"L_code"].split(","):
                    if semester_code.get() in df_sel.at[index_new,"semL"].split(","):
                        if assign_dept_code.get() in df_sel.at[index_new,"deptL_class"].split(","):
                            code_index = df_sel.at[index_new,"L_code"].split(",").index(sub_code.get())
                            #df_sel.loc[index_new,"L_hr"] = int(df_sel.at[index_new,"L_hr"])+1
                            df_sel.loc[index_new,"L_hr"][code_index] = int(df_sel.at[0,"L_hr"]) if code_index==0 else int(df_sel.at[0,"L_hr"][code_index])+1
                            df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
                else:    
                    df_sel.loc[index_new,"deptL_class"] = f"{df_sel.at[index_new,"deptL_class"]},{assign_dept_code.get()}"
                    df_sel.loc[index_new,"semL"] = f"{df_sel.at[index_new,"semL"]},{semester_code.get()}"
                    df_sel.loc[index_new,"L_code"] = f"{df_sel.at[index_new,"L_code"]},{sub_code.get()}"
                    df_sel.loc[index_new,"L_class"] = f"{df_sel.at[index_new,"L_class"]},{lec_class_hr}"
                    df_sel.loc[index_new,"L_hr"] = f"{df_sel.at[index_new,"L_hr"]},{1}"
                    df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
            else:
                new_record = pd.DataFrame([{'nick_name':faculty_name,'emp_code':dept_emp_code.get(),'dept_origin':dept_code.get(),
                    'deptL_class':assign_dept_code.get(),'semL':semester_code.get(),'L_code':sub_code.get(),'L_class':lec_class_hr,"L_hr":1}])
                df_sel = pd.concat([df_sel,new_record],ignore_index=True)    
                df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
            print(df_sel)
        case "Tutorial":
            if faculty_name in list(df_sel.nick_name):
                index_new = list(df_sel.nick_name).index(faculty_name)
                print(index_new)
                if sub_code.get() in df_sel.at[index_new,"T_code"].split(","):
                    if semester_code.get() in df_sel.at[index_new,"semT"].split(","):
                        if assign_dept_code.get() not in df_sel.at[index_new,"deptT_class"].split(","):
                            df_sel.loc[index_new,"deptT_class"] = f"{df_sel.at[index_new,"deptT_class"]},{assign_dept_code.get()}"
                            df_sel.loc[index_new,"semT"] = f"{df_sel.at[index_new,"semT"]},{semester_code.get()}"
                            df_sel.loc[index_new,"T_code"] = f"{df_sel.at[index_new,"T_code"]},{sub_code.get()}"
                            df_sel.loc[index_new,"T_hr"] = f"{df_sel.at[index_new,"T_hr"]},{1}"
                            df_sel.to_excel('output_data.xlsx',dept_code.get(),index=False,engine="openpyxl")
            else:
                new_record = pd.DataFrame([{'nick_name':faculty_name,'emp_code':dept_emp_code.get(),'dept_origin':dept_code.get(),
                    'deptT_class':assign_dept_code.get(),'semT':semester_code.get(),'T_code':sub_code.get(),"T_hr":1}])
                df_sel = pd.concat([df_sel,new_record],ignore_index=True)    
                df_sel.to_excel('output_data.xlsx',dept_code.get(),index=False,engine="openpyxl")
        case "Tutorial":
            if faculty_name in list(df_sel.nick_name):
                index_new = list(df_sel.nick_name).index(faculty_name)
                print(index_new)
                if sub_code.get() in df_sel.at[index_new,"P_code"].split(","):
                    if semester_code.get() in df_sel.at[index_new,"semP"].split(","):
                        if assign_dept_code.get() not in df_sel.at[index_new,"deptP_class"].split(","):
                            df_sel.loc[index_new,"deptP_class"] = f"{df_sel.at[index_new,"deptP_class"]},{assign_dept_code.get()}"
                            df_sel.loc[index_new,"semP"] = f"{df_sel.at[index_new,"semP"]},{semester_code.get()}"
                            df_sel.loc[index_new,"P_code"] = f"{df_sel.at[index_new,"P_code"]},{sub_code.get()}"
                            df_sel.loc[index_new,"P_hr"] = f"{df_sel.at[index_new,"P_hr"]},{1}"
                            df_sel.to_excel('output_data.xlsx',dept_code.get(),index=False,engine="openpyxl")
            else:
                new_record = pd.DataFrame([{'nick_name':faculty_name,'emp_code':dept_emp_code.get(),'dept_origin':dept_code.get(),
                    'deptP_class':assign_dept_code.get(),'semP':semester_code.get(),'P_code':sub_code.get(),"P_hr":2}])
                df_sel = pd.concat([df_sel,new_record],ignore_index=True)    
                df_sel.to_excel('output_data.xlsx',dept_code.get(),index=False,engine="openpyxl")             
def cancel_option_selection(event):
    lf1_lf2_clear()
    pass
def lf1_lf2_clear():
    dept_emp_code.set('')
    dept_emp_code['values']=()
    assign_dept_code.set('')
    odd_even_code.set('')
    semester_code.set('')
    ltp_code.set('')
    sub_code.set('')
    cancel_code.set('')

def common_for_cancel_options():
    pass

def clear_day_slot():
    text_box1.delete('1.0','end')
    text_box2.delete('1.0','end')
    sr0.set(0)
    sr1.set(0)
    sr2.set(0)
    sr3.set(0)
    sr4.set(0)
    sr5.set(0)
    sr6.set(0)

def cancel_option(event):
    global clear_status
    cancel_entry = cancel_click.get()
    if cancel_entry == "Assigned Class":
        clear_status = 1
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")    
    elif cancel_entry == "Day & Slot":
        clear_status = 2
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")
    else:  
        clear_status = 3
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")
def confirm_cancel():
    pass

root=Tk()
root.title("Data entry by Faculty for TimeTable")
w_width = root.winfo_screenwidth()
w_height = root.winfo_screenheight()
dev_WIDTH = 700
dev_HEIGHT = 550
ref_s_WIDTH = 1920
ref_s_HEIGHT = 1080
s_width = int(dev_WIDTH*w_width/ref_s_WIDTH)
s_height = int(dev_HEIGHT*w_height/ref_s_HEIGHT)
#print(s_width,s_height)
root.geometry("700x550")
root.iconbitmap('Logo.ico')
root.columnconfigure(0,weight=1) #weight= 1 indicates scale of one
root.rowconfigure(0,weight=1)
frame=Frame(root)
frame.grid(row=0,column=0,padx=10,pady=5,sticky="nsew")
x = "Lecture"
global data_tt,my_data_tt,in_tt
in_tt={"week_day":" ","slot_1":[],"slot_2":[],"slot_3":[],
    "slot_4":[],"slot_5":[],"slot_6":[],"slot_x":[]}
odd_even="odd"
day_slot={"week_slot":3,"time_slot0":0,"time_slot1":1,"time_slot2":2,"time_slot3":3,
    "time_slot4":4,"time_slot5":5,"time_slot6":6}
sr=IntVar()
#Storage file for manipulation
data_ref = [{'nick_name':'','emp_code':'','dept_origin':'','deptL_class':'','semL':'','L_code':'','L_class':'',"L_hr":'',
    'deptT_class':'','semT':'','T_code':'',"T_hr":'','deptP_class':'','semP':'','P_code':'',"P_hr":'',
    'deptR_class':'','semR':'','R_code':'',"R_hr":''}]
#Storage file on departments
dept_data = [{'emp_code':'','dept':'','grade':'','nick_name':'','pay_band':''}]
#Storage file on curiculam
curi_data = [{'Code':'','Subject':'','Select':'','Number':'','L':'','T':'','P':'','R':''}]

#LBELFRAME lf1
frame.columnconfigure(0,weight=1) #weight= 1 indicates scale of one
frame.rowconfigure((0,1,2,3),weight=1)
lf1=LabelFrame(frame,text="Employ Code: Parent Dept: Faculty Initial: Assigned Dept.",padx=5,pady=2,fg="Blue")
lf1.grid(row=0,column=0,padx=10,pady=5,sticky="nsew")
lf1.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf1.rowconfigure((0,1),weight=1)
dept_emp_code=Label(lf1,text="Parent Dept.& Emp.Code",width=15)
dept_emp_code.grid(row=0,column=0,columnspan=2,padx=2,pady=2,sticky="nsew")
dept_class=Label(lf1,text="Assigned Department/Semester",width=15)
dept_class.grid(row=0,column=2,columnspan=2,padx=2,pady=2,sticky="nsew")
sub_lec_faculty = {}
sub_tut_faculty = {}
sub_lab_faculty = {}
sub_proj_faculty = {}
#lf1 entries
dept_options=StringVar()
dept_code=ttk.Combobox(lf1,textvariable=dept_options,state='raedonly')
dept_code.grid(row=1,column=0,padx=10,pady=2,sticky="nsew")
dept_code['values']=("CE","ME","EE")
dept_code.current()
dept_code.bind("<<ComboboxSelected>>",parent_dept_selection)
dept_emp_options=StringVar()
dept_emp_code=ttk.Combobox(lf1,textvariable=dept_emp_options,state='raedonly')
dept_emp_code.grid(row=1,column=1,padx=10,pady=2,sticky="nsew")
dept_emp_code['values']=()
dept_emp_code.current()
dept_emp_code.bind("<<ComboboxSelected>>",emp_code_selection)
assign_dept_options=StringVar()
assign_dept_code=ttk.Combobox(lf1,textvariable=assign_dept_options,state='raedonly')
assign_dept_code.grid(row=1,column=2,padx=10,pady=2,sticky="nsew")
assign_dept_code['values']=()
assign_dept_code.current()
assign_dept_code.bind("<<ComboboxSelected>>",dept_option_selection)
odd_even_options=StringVar()
odd_even_code=ttk.Combobox(lf1,textvariable=odd_even_options,state='raedonly')
odd_even_code.grid(row=1,column=3,padx=10,pady=2,sticky="nsew")
odd_even_code['values']=()
odd_even_code.current()
odd_even_code.bind("<<ComboboxSelected>>",odd_even_selection)
#LBELFRAME lf2
lf2=LabelFrame(frame,text="Selecting Assigned Classes",padx=5,pady=2,fg="Blue")
lf2.grid(row=1,column=0,padx=10,pady=5,sticky="nsew")
#lf2 options
global semester_options
#lf2 labels
lf2.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf2.rowconfigure((0,1,2),weight=1)
text_box1=Text(lf2, width=60,height=3,fg="red",font=("Ariel",11))
text_box1.grid(row=0, column=0,columnspan=4,padx=10,pady=5,sticky="nsew")
semester_class=Label(lf2,text="Semester",width=15)
semester_class.grid(row=1,column=0,padx=2,pady=2,sticky="nsew")
ltp_type=Label(lf2,text="Class type",width=15)
ltp_type.grid(row=1,column=1,padx=2,pady=2,sticky="nsew")
sub_code=Label(lf2,text="Subject Code",width=15)
sub_code.grid(row=1,column=2,padx=2,pady=2,sticky="nsew")
sub_check=Label(lf2,text="ACTION",width=15)
sub_check.grid(row=1,column=3,padx=2,pady=2,sticky="nsew")
#lf2 entries
semester_options=StringVar()
semester_code=ttk.Combobox(lf2,textvariable=semester_options,state='raedonly')
semester_code.grid(row=2,column=0,padx=10,pady=2,sticky="nsew")
semester_code['values']=()
semester_code.current()
semester_code.bind("<<ComboboxSelected>>",semester_option_selection)
ltp_options=StringVar()
ltp_code=ttk.Combobox(lf2,textvariable=ltp_options,state='raedonly')
ltp_code.grid(row=2,column=1,padx=10,pady=2,sticky="nsew")
ltp_code['values']=()
ltp_code.current()
ltp_code.bind("<<ComboboxSelected>>",ltp_option_selection)
sub_options=StringVar()
sub_code=ttk.Combobox(lf2,textvariable=sub_options,state='raedonly')
sub_code.grid(row=2,column=2,padx=10,pady=2,sticky="nsew")
sub_code['values']=()
sub_code.current()
sub_code.bind("<<ComboboxSelected>>",sub_option_selection)
action_options=StringVar()
action_code=ttk.Combobox(lf2,textvariable=action_options,state='raedonly')
action_code.grid(row=2,column=3,padx=10,pady=2,sticky="nsew")
action_code['values']=("CLEAR")
action_code.current()
action_code.bind("<<ComboboxSelected>>",action_option_selection)
#LBELFRAME lf3
lf3=LabelFrame(frame,text="Timetable Weekday and Slot selection",padx=5,pady=2,fg="Blue")
lf3.grid(row=2,column=0,padx=10,pady=5,sticky="nsew")
lf3.columnconfigure((0,1,2,3,4,5,6),weight=1,uniform="a") #weight= 1 indicates scale of one
lf3.rowconfigure((0,1,2),weight=1)
r3=IntVar()
week_day=["Monday","Tuesday","Wednesday","Thursday","Friday"]
for i in range(len(week_day)):
    week=(Radiobutton(lf3,text=week_day[i],variable=r3,value=i,command=lambda:week_click(r3.get())))
    week.grid(row=0,column=i+1,padx=10,pady=2,sticky="nsew")
refresh_button = Button(lf3,text="REFRESH",command=refresh_weekslot)
refresh_button.grid(row=0,column=6,padx=2,pady=2,sticky="nsew")
#print(week.r3)
show_msg=Label(lf3,text="-----------Slot 1 to Slot 6 for Lecture & Tutorial and Slot X optional-------------",fg="Blue")
show_msg.grid(row=1, column=0, columnspan=7,padx=2,pady=2,sticky="nsew")
slots=["Slot 1","Slot 2","Slot 3","Slot 4","Slot 5","Slot 6","Slot X"]
sr0 = IntVar()
sr1 = IntVar()
sr2 = IntVar()
sr3 = IntVar()
sr4 = IntVar()
sr5 = IntVar()
sr6 = IntVar()
global cb1
cb1=Checkbutton(lf3, text=slots[0], variable=sr0,state='disabled')
cb1.grid(row=3, column=0,sticky="nsew")
cb2=Checkbutton(lf3, text=slots[1], variable=sr1,state='disabled')
cb2.grid(row=3, column=1,sticky="nsew")
cb3=Checkbutton(lf3, text=slots[2], variable=sr2,state='disabled')
cb3.grid(row=3, column=2,sticky="nsew")
cb4=Checkbutton(lf3, text=slots[3], variable=sr3,state='disabled')
cb4.grid(row=3, column=3,sticky="nsew")
cb5=Checkbutton(lf3, text=slots[4], variable=sr4,state='disabled')
cb5.grid(row=3, column=4,sticky="nsew")
cb6=Checkbutton(lf3, text=slots[5], variable=sr5,state='disabled')
cb6.grid(row=3, column=5,sticky="nsew")
cb7=Checkbutton(lf3, text=slots[6], variable=sr6,state='disabled')
cb7.grid(row=3, column=6,sticky="nsew")
#LBELFRAME lf4
#lf4 options
lf4=LabelFrame(frame,text="Timetable Data Entry Commands",padx=5,pady=2,fg="Blue")
lf4.grid(row=3,column=0,padx=10,pady=5,sticky="nsew")
lf4.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf4.rowconfigure((0,1),weight=1)
text_box2=Text(lf4, width=60,height=3,fg="red",font=("Ariel",11))
text_box2.grid(row=0,column=0,columnspan=4,padx=10,pady=5,sticky="nsew")
cancel_check=Label(lf4,text="Cancel Data Entry",width=15)
cancel_check.grid(row=1,column=0,columnspan=2, padx=2,pady=2,sticky="nsew")
cancel_options=StringVar()
cancel_code=ttk.Combobox(lf4,textvariable=cancel_options,state='raedonly')
cancel_code.grid(row=2,column=0,padx=10,pady=2,sticky="nsew")
cancel_code['values']=("Clear Day & Slot ","Clear All Data","Clear Data from Time Table")
cancel_code.current()
cancel_code.bind("<<ComboboxSelected>>",cancel_option_selection)

cancel_button=Button(lf4,text="Confirm CANCEL",bg="white",command=confirm_cancel)
cancel_button.grid(row=2,column=1,padx=2,pady=2,sticky="nsew")
verify_data=Label(lf4,text="Verify & Update TimeTable",width=15)
verify_data.grid(row=1,column=2,columnspan=2, padx=2,pady=2,sticky="nsew")
verify_button=Button(lf4,text="Verified OK",bg="white",command=verify_ok)
verify_button.grid(row=2,column=2,padx=2,pady=2,sticky="nsew")
update_button=Button(lf4,text="Update TimeTable",bg="white",command=update_tt)
update_button.grid(row=2,column=3,padx=2,pady=2,sticky="nsew")
root.mainloop()
#