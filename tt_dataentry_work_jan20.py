# Timetable project
# Intermediate Stage
# Data Entry upto lf#1, lf#2 complete
from tkinter import*
from tkinter import ttk
import tkinter as tk
from openpyxl import load_workbook
import pandas as pd
from tt_show_class_jan20 import TimeTable

def find_key(dictionary,value):
    for key,val in dictionary.items():
        if val == value:
            return(key)
    return None
def excel_read1(file_name,sheet_name):
    data_dfrow = []
    lst_dfcol = []
    workbook = load_workbook(filename=f"{file_name}.xlsx")
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rowdf_dict = dict(zip(headers, row))
        data_dfrow.append(rowdf_dict)
    for column in sheet.iter_cols(min_row=2, values_only=True):
        df_col = []
        for cell in column:
            df_col.append(cell)
        lst_dfcol.append(df_col)
        coldf_dict = dict(zip(headers, lst_dfcol))
    workbook.close()
    row_df_ret = [data_dfrow,coldf_dict,headers]
    return row_df_ret
def parent_dept_selection(event): 
    global data_faculty
    action_code.set('')
    action_code.config(values=(''))
    emp_code.set('')
    data_faculty = excel_read1("faculty_data",dept_code.get())
    emp_code.config(values=data_faculty[1]["emp_code"])
def emp_code_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    assign_dept_code.set('')
    assign_dept_code.config(values=("CE","ME","EE"))
def dept_option_selection(event):
    odd_even_code.config(values=("ODD","EVEN"))
    action_code.set('')
    action_code.config(values=('')) 
def odd_even_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    odd_even = odd_even_code.get()
    if odd_even == 'ODD':
        odd_even_list = ['s1','s3','s5','s7']
    elif odd_even == 'EVEN':
        odd_even_list = ['s2','s4','s6','s8']
    else:
        pass
    semester_code.set('')
    semester_code.config(values=odd_even_list)
def semester_option_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    ltp_code.set('')
    ltp_code.config(values=("Lecture","Tutorial","Practical","Projects"))
def ltp_option_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    global data_faculty,data_tt,data_temp,faculty_sel
    data_curi = excel_read1(f"curiculam_{assign_dept_code.get()}",semester_code.get())
    new_row_lst = [data_curi[0][i] for i in range(len(data_curi[1]["Select"])) if data_curi[1]["Select"][i] != 0]
    data_col_lst = list(data_curi[1].values())  
    cur_col_lst = [[data_col_lst[j][i] for i in range(len(data_curi[1]["Select"])) if data_curi[1]["Select"][i] != 0] for j in range(len(data_curi[2]))]
    cur_col_dict = dict(zip(data_curi[2], cur_col_lst))
    data_tt = excel_read1(f"timeTable_{assign_dept_code.get()}",semester_code.get())
    data_temp = excel_read1(f"faculty_assignment_{odd_even_code.get()}",assign_dept_code.get())
    faculty_sel = (data_faculty[1]["nick_name"][(data_faculty[1]["emp_code"].index(int(emp_code.get())))]) #from faculty file
    match ltp_code.get():
        case "Lecture":
            cur_lec_lst = [(cur_col_dict["Code"])[i] for i in range(len(cur_col_dict["Code"])) if list(cur_col_dict["L"])[i] != 0]
            cur_L_lst = [(cur_col_dict["L"])[i] for i in range(len(cur_col_dict["L"])) if list(cur_col_dict["L"])[i] != 0]
        # avoid duplicate entry in Lecture
            new_lec_lst = [lec for lec in cur_lec_lst if cur_L_lst[cur_lec_lst.index(lec)] > len([i for j in range(1,len(data_tt[2])) for i in range(len(data_tt[1][data_tt[2][j]])) if data_tt[1][data_tt[2][j]][i] == lec])]
            code_lst = []
            for i in range(len(new_lec_lst)):
                if new_lec_lst[i] not in data_temp[1]["L_code"]:
                    code_lst.append(new_lec_lst[i])
                else:
                    if faculty_sel == data_temp[1]["nick_name"][data_temp[1]["L_code"].index(new_lec_lst[i])]:
                        code_lst.append(new_lec_lst[i])
                        print(faculty_sel,data_temp[1]["nick_name"][(data_temp[1]["L_code"].index(new_lec_lst[i]))])             
        case "Tutorial":
            cur_tut_lst = [(cur_col_dict["Code"])[i] for i in range(len(cur_col_dict["Code"])) if list(cur_col_dict["T"])[i] != 0]
            cur_T_lst = [(cur_col_dict["T"])[i] for i in range(len(cur_col_dict["T"])) if list(cur_col_dict["T"])[i] != 0]
            cur_num_lst = [(cur_col_dict["Number"])[i] for i in range(len(cur_col_dict["Number"])) if list(cur_col_dict["T"])[i] != 0]
            new_tut_lst = [tut for tut in cur_tut_lst if cur_T_lst[cur_tut_lst.index(tut)] > len([i for j in range(1,len(data_tt[2])) for i in range(len(data_tt[1][data_tt[2][j]])) if data_tt[1][data_tt[2][j]][i] == tut])]
            """    
        # count faculty for tutorial    
            T_num_lst = []
            for i in range(len(cur_tut_lst)):
                if cur_num_lst[i] <= 26:
                    T_count = 1
                elif cur_num_lst[i] <= 48:
                    T_count = 2
                else:
                    T_count = 3
                T_num_lst.append(T_count)
            print(T_num_lst)
            """
            code_lst = []
            for i in range(len(cur_tut_lst)):
                if cur_tut_lst[i] not in data_temp[1]["T_code"]:
                    code_lst.append(cur_tut_lst[i])
                else:
                    if faculty_sel != data_temp[1]["nick_name"][data_temp[1]["T_code"].index(cur_tut_lst[i])]:
                        code_lst.append(cur_tut_lst[i])              
        case "Practical":
            cur_lab_lst = [(cur_col_dict["Code"])[i] for i in range(len(cur_col_dict["Code"])) if list(cur_col_dict["P"])[i] != 0]
            cur_P_lst = [(cur_col_dict["P"])[i] for i in range(len(cur_col_dict["P"])) if list(cur_col_dict["P"])[i] != 0]
            cur_num_lst = [(cur_col_dict["Number"])[i] for i in range(len(cur_col_dict["Number"])) if list(cur_col_dict["P"])[i] != 0]
            new_lab_lst = [lab for lab in cur_lab_lst if cur_P_lst[cur_lab_lst.index(lab)] > len([i for j in range(1,len(data_tt[2])) for i in range(len(data_tt[1][data_tt[2][j]])) if data_tt[1][data_tt[2][j]][i] == lab])]          
            code_lst = []
            for i in range(len(cur_lab_lst)):
                if cur_lab_lst[i] not in data_temp[1]["P_code"]:
                    code_lst.append(cur_lab_lst[i])
                else:
                    if faculty_sel != data_temp[1]["nick_name"][data_temp[1]["T_code"].index(cur_lab_lst[i])]:
                        code_lst.append(cur_lab_lst[i])                 
        case "Projects":
            cur_pro_lst = [(cur_col_dict["Code"])[i] for i in range(len(cur_col_dict["Code"])) if list(cur_col_dict["R"])[i] != 0]
            cur_R_lst = [(cur_col_dict["R"])[i] for i in range(len(cur_col_dict["R"])) if list(cur_col_dict["R"])[i] != 0]
            cur_num_lst = [(cur_col_dict["Number"])[i] for i in range(len(cur_col_dict["Number"])) if list(cur_col_dict["R"])[i] != 0]
            new_pro_lst = [pro for pro in cur_pro_lst if cur_R_lst[cur_pro_lst.index(pro)] > len([i for j in range(1,len(data_tt[2])) for i in range(len(data_tt[1][data_tt[2][j]])) if data_tt[1][data_tt[2][j]][i] == pro])]
            code_lst = new_pro_lst
        case _:
            print("Who")
    sub_code.set('')
    sub_code.config(values=code_lst)
def sub_option_selection(event):
    #data_tt["sub_code"] = sub_code.get()
    action_code.config(values=("PROCEED","CLEAR"))
def action_option_selection(event):
    global data_tt,data_faculty
    global class_x,data,ro_co_dict,row_val
    action_entry=action_code.get()
    #print(data_tt)
    if action_entry == "CLEAR":
        lf1_lf2_clear()
        pass
    elif action_entry == "PROCEED":
        TimeTable(root,f"timeTable_{assign_dept_code.get()}.xlsx",semester_code.get())
        ro_co_dict = {}
        for i in range(5):
            ro_co_dict.update({i:[0 for _ in range(7)]})
        week_click(None)
        refresh_button.config(bg="Blue",fg="white")
        text_box2.delete('1.0','end')
        text_box2.insert('1.0',f"Select weekday and one from available slot before REFRESH")
def week_click(week_value:any):
    cb1.config(state='disabled')
    cb2.config(state='disabled')
    cb3.config(state='disabled')
    cb4.config(state='disabled')
    cb5.config(state='disabled')
    cb6.config(state='disabled')
    cb7.config(state='disabled')
    sr0.set(0)
    sr1.set(0)
    sr2.set(0)
    sr3.set(0)
    sr4.set(0)
    sr5.set(0)
    sr6.set(0)
    if data_tt[0][r3.get()][data_tt[2][1]] == "FREE":
        cb1.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][2]] == "FREE":
        cb2.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][3]] == "FREE":
        cb3.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][4]] == "FREE":
        cb4.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][5]] == "FREE":
        cb5.config(state='normal')
    if data_tt[0][r3.get()][data_tt[2][6]] == "FREE":
        cb6.config(state='normal')
    verify_button.config(bg="white",fg="black")
def refresh_weekslot():
    ro_co_dict[r3.get()] = [sr0.get(),sr1.get(),sr2.get(),sr3.get(),sr4.get(),sr5.get(),sr6.get()]
    if 1 not in ro_co_dict[r3.get()]:
        text_box2.insert('1.0',f"!!WARNING!! Select weekday and one from available slot before REFRESH...................")
        return
    #print(r3.get(),ro_co_dict[r3.get()].index(1))
    refresh_button.config(bg="yellow",fg="black")
    verify_button.config(bg="Blue",fg="white")
    update_button.config(bg="white",fg="black")
    text_box2.delete('1.0','end')
    text_box2.insert('1.0',f"Confirm your choice {ltp_code.get()} : for the subject {sub_code.get()} for {semester_code.get()} in {assign_dept_code.get()} .... ")
    text_box2.insert('2.0',f"Selected choice {data_tt[1]["week_day"][r3.get()]} for the slot {data_tt[2][(ro_co_dict[r3.get()].index(1))+1]} ")
def verify_ok():
    update_button.config(bg="Blue",fg="white")
    verify_button.config(bg="white",fg="black")
    refresh_button.config(bg="white",fg="black")
    text_box2.delete('1.0','end')
    text_box2.insert('1.0',f"Confirm and SAVE to Time Table")
def update_tt():
    data_tt = excel_read1(f"timeTable_{assign_dept_code.get()}",semester_code.get())
    data_faculty = excel_read1("faculty_data",dept_code.get())
    data_temp = excel_read1(f"faculty_assignment_{odd_even_code.get()}",assign_dept_code.get())
    #print(data_temp[1])
    #print()
    #print(data_temp[2])
    #print(data_faculty[2])
    #print()
    update_button.config(bg="white",fg="black")
    #print(data_faculty[1]["emp_code"].index(int(emp_code.get())))
    faculty_ref = data_faculty[1]["nick_name"][data_faculty[1]["emp_code"].index(int(emp_code.get()))]
    #print(faculty_ref)
    match ltp_code.get():
        case "Lecture":
            if faculty_ref in data_temp[1]["nick_name"]:
                temp_rowid = data_temp[1]["nick_name"].index(faculty_ref)
                print(faculty_ref,temp_rowid,data_temp[1]["nick_name"][data_temp[1]["nick_name"].index(faculty_ref)])
                print(data_temp[2])
                print(data_temp[1]['emp_code'][temp_rowid])
            else:
                print("NO")
    return

    match ltp_code.get():
        case "Lecture":
            wb = load_workbook(f"faculty_assignment_{odd_even_code.get()}.xlsx")
            sheet = wb[assign_dept_code.get()]
            headers = [cell.value for cell in sheet[1]]
            print(headers)
            nick_name_lst = []
            code_select_lst = []
            for i in range(len(sheet['A'])-1):
                nick_name_lst.append(sheet.cell(row=i+2,column=1).value)
                code_select_lst.append(sheet.cell(row=i+2,column=5).value)
            print(nick_name_lst)
            print(code_select_lst)

    #@@@@@@@@@@@@@@@
    text_box2.delete('1.0','end')
    text_box2.insert('1.0',f"Timetable updated for {assign_dept_code.get()} : {semester_code.get()} :  Lecture {sub_code.get()}")
    TimeTable(root,f"timeTable_{assign_dept_code.get()}.xlsx",semester_code.get())
    text_box2.delete('1.0','end')
    if lec_count != lec_intt:
        text_box2.insert('1.0',f"{sub_code.get()} : {lec_count} {ltp_code.get()} slots, {lec_intt} saved. ENTER all slots for {sub_code.get()}")
    else:
        text_box2.insert('1.0',f"{sub_code.get()} : All {lec_count} {ltp_code.get()} slots saved in Timetable")
def cancel_option_selection(event):
    lf1_lf2_clear()
    pass
def lf1_lf2_clear():
    emp_code.set('')
    emp_code['values']=()
    assign_dept_code.set('')
    odd_even_code.set('')
    semester_code.set('')
    ltp_code.set('')
    sub_code.set('')
    cancel_code.set('')

def common_for_cancel_options():
    pass

def clear_day_slot():
    text_box1.delete('1.0','end')
    text_box2.delete('1.0','end')
    sr0.set(0)
    sr1.set(0)
    sr2.set(0)
    sr3.set(0)
    sr4.set(0)
    sr5.set(0)
    sr6.set(0)

def cancel_option(event):
    global clear_status
    cancel_entry = cancel_click.get()
    if cancel_entry == "Assigned Class":
        clear_status = 1
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")    
    elif cancel_entry == "Day & Slot":
        clear_status = 2
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")
    else:  
        clear_status = 3
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")
def confirm_cancel():
    pass

root=Tk()
root.title("Data entry by Faculty for TimeTable")
w_width = root.winfo_screenwidth()
w_height = root.winfo_screenheight()
dev_WIDTH = 700
dev_HEIGHT = 550
ref_s_WIDTH = 1920
ref_s_HEIGHT = 1080
s_width = int(dev_WIDTH*w_width/ref_s_WIDTH)
s_height = int(dev_HEIGHT*w_height/ref_s_HEIGHT)
#print(s_width,s_height)
root.geometry("700x550")
root.iconbitmap('Logo.ico')
root.columnconfigure(0,weight=1) #weight= 1 indicates scale of one
root.rowconfigure(0,weight=1)
frame=Frame(root)
frame.grid(row=0,column=0,padx=10,pady=5,sticky="nsew")
x = "Lecture"
global data_tt,my_data_tt,in_tt
in_tt={"week_day":" ","slot_1":[],"slot_2":[],"slot_3":[],
    "slot_4":[],"slot_5":[],"slot_6":[],"slot_x":[]}
odd_even="odd"
day_slot={"week_slot":3,"time_slot0":0,"time_slot1":1,"time_slot2":2,"time_slot3":3,
    "time_slot4":4,"time_slot5":5,"time_slot6":6}
sr=IntVar()
#LBELFRAME lf1
frame.columnconfigure(0,weight=1) #weight= 1 indicates scale of one
frame.rowconfigure((0,1,2,3),weight=1)
lf1=LabelFrame(frame,text="Employ Code: Parent Dept: Faculty Initial: Assigned Dept.",padx=5,pady=2,fg="Blue")
lf1.grid(row=0,column=0,padx=10,pady=5,sticky="nsew")
lf1.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf1.rowconfigure((0,1),weight=1)
dept_emp_code=Label(lf1,text="Parent Dept.& Emp.Code",width=15)
dept_emp_code.grid(row=0,column=0,columnspan=2,padx=2,pady=2,sticky="nsew")
dept_class=Label(lf1,text="Assigned Department/Semester",width=15)
dept_class.grid(row=0,column=2,columnspan=2,padx=2,pady=2,sticky="nsew")
sub_lec_faculty = {}
sub_tut_faculty = {}
sub_lab_faculty = {}
sub_proj_faculty = {}
#lf1 entries
dept_options=StringVar()
dept_code=ttk.Combobox(lf1,textvariable=dept_options,state='raedonly')
dept_code.grid(row=1,column=0,padx=10,pady=2,sticky="nsew")
dept_code['values']=("CE","ME","EE")
dept_code.current()
dept_code.bind("<<ComboboxSelected>>",parent_dept_selection)
emp_options=StringVar()
emp_code=ttk.Combobox(lf1,textvariable=emp_options,state='raedonly')
emp_code.grid(row=1,column=1,padx=10,pady=2,sticky="nsew")
emp_code['values']=()
emp_code.current()
emp_code.bind("<<ComboboxSelected>>",emp_code_selection)
assign_dept_options=StringVar()
assign_dept_code=ttk.Combobox(lf1,textvariable=assign_dept_options,state='raedonly')
assign_dept_code.grid(row=1,column=2,padx=10,pady=2,sticky="nsew")
assign_dept_code['values']=()
assign_dept_code.current()
assign_dept_code.bind("<<ComboboxSelected>>",dept_option_selection)
odd_even_options=StringVar()
odd_even_code=ttk.Combobox(lf1,textvariable=odd_even_options,state='raedonly')
odd_even_code.grid(row=1,column=3,padx=10,pady=2,sticky="nsew")
odd_even_code['values']=()
odd_even_code.current()
odd_even_code.bind("<<ComboboxSelected>>",odd_even_selection)
#LBELFRAME lf2
lf2=LabelFrame(frame,text="Selecting Assigned Classes",padx=5,pady=2,fg="Blue")
lf2.grid(row=1,column=0,padx=10,pady=5,sticky="nsew")
#lf2 options
global semester_options
#lf2 labels
lf2.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf2.rowconfigure((0,1,2),weight=1)
text_box1=Text(lf2, width=60,height=3,fg="red",font=("Ariel",11))
text_box1.grid(row=0, column=0,columnspan=4,padx=10,pady=5,sticky="nsew")
semester_class=Label(lf2,text="Semester",width=15)
semester_class.grid(row=1,column=0,padx=2,pady=2,sticky="nsew")
ltp_type=Label(lf2,text="Class type",width=15)
ltp_type.grid(row=1,column=1,padx=2,pady=2,sticky="nsew")
sub_code=Label(lf2,text="Subject Code",width=15)
sub_code.grid(row=1,column=2,padx=2,pady=2,sticky="nsew")
sub_check=Label(lf2,text="ACTION",width=15)
sub_check.grid(row=1,column=3,padx=2,pady=2,sticky="nsew")
#lf2 entries
semester_options=StringVar()
semester_code=ttk.Combobox(lf2,textvariable=semester_options,state='raedonly')
semester_code.grid(row=2,column=0,padx=10,pady=2,sticky="nsew")
semester_code['values']=()
semester_code.current()
semester_code.bind("<<ComboboxSelected>>",semester_option_selection)
ltp_options=StringVar()
ltp_code=ttk.Combobox(lf2,textvariable=ltp_options,state='raedonly')
ltp_code.grid(row=2,column=1,padx=10,pady=2,sticky="nsew")
ltp_code['values']=()
ltp_code.current()
ltp_code.bind("<<ComboboxSelected>>",ltp_option_selection)
sub_options=StringVar()
sub_code=ttk.Combobox(lf2,textvariable=sub_options,state='raedonly')
sub_code.grid(row=2,column=2,padx=10,pady=2,sticky="nsew")
sub_code['values']=()
sub_code.current()
sub_code.bind("<<ComboboxSelected>>",sub_option_selection)
action_options=StringVar()
action_code=ttk.Combobox(lf2,textvariable=action_options,state='raedonly')
action_code.grid(row=2,column=3,padx=10,pady=2,sticky="nsew")
action_code['values']=("CLEAR")
action_code.current()
action_code.bind("<<ComboboxSelected>>",action_option_selection)
#LBELFRAME lf3
lf3=LabelFrame(frame,text="Timetable Weekday and Slot selection",padx=5,pady=2,fg="Blue")
lf3.grid(row=2,column=0,padx=10,pady=5,sticky="nsew")
lf3.columnconfigure((0,1,2,3,4,5,6),weight=1,uniform="a") #weight= 1 indicates scale of one
lf3.rowconfigure((0,1,2),weight=1)
r3=IntVar()
week_day=["Monday","Tuesday","Wednesday","Thursday","Friday"]
for i in range(len(week_day)):
    week=(Radiobutton(lf3,text=week_day[i],variable=r3,value=i,command=lambda:week_click(r3.get())))
    week.grid(row=0,column=i+1,padx=10,pady=2,sticky="nsew")
refresh_button = Button(lf3,text="REFRESH",command=refresh_weekslot)
refresh_button.grid(row=0,column=6,padx=2,pady=2,sticky="nsew")
#print(week.r3)
show_msg=Label(lf3,text="-----------Slot 1 to Slot 6 for Lecture & Tutorial and Slot X optional-------------",fg="Blue")
show_msg.grid(row=1, column=0, columnspan=7,padx=2,pady=2,sticky="nsew")
slots=["Slot 1","Slot 2","Slot 3","Slot 4","Slot 5","Slot 6","Slot X"]
sr0 = IntVar()
sr1 = IntVar()
sr2 = IntVar()
sr3 = IntVar()
sr4 = IntVar()
sr5 = IntVar()
sr6 = IntVar()
global cb1
cb1=Checkbutton(lf3, text=slots[0], variable=sr0,state='disabled')
cb1.grid(row=3, column=0,sticky="nsew")
cb2=Checkbutton(lf3, text=slots[1], variable=sr1,state='disabled')
cb2.grid(row=3, column=1,sticky="nsew")
cb3=Checkbutton(lf3, text=slots[2], variable=sr2,state='disabled')
cb3.grid(row=3, column=2,sticky="nsew")
cb4=Checkbutton(lf3, text=slots[3], variable=sr3,state='disabled')
cb4.grid(row=3, column=3,sticky="nsew")
cb5=Checkbutton(lf3, text=slots[4], variable=sr4,state='disabled')
cb5.grid(row=3, column=4,sticky="nsew")
cb6=Checkbutton(lf3, text=slots[5], variable=sr5,state='disabled')
cb6.grid(row=3, column=5,sticky="nsew")
cb7=Checkbutton(lf3, text=slots[6], variable=sr6,state='disabled')
cb7.grid(row=3, column=6,sticky="nsew")
#LBELFRAME lf4
#lf4 options
lf4=LabelFrame(frame,text="Timetable Data Entry Commands",padx=5,pady=2,fg="Blue")
lf4.grid(row=3,column=0,padx=10,pady=5,sticky="nsew")
lf4.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf4.rowconfigure((0,1),weight=1)
text_box2=Text(lf4, width=60,height=3,fg="red",font=("Ariel",11))
text_box2.grid(row=0,column=0,columnspan=4,padx=10,pady=5,sticky="nsew")
cancel_check=Label(lf4,text="Cancel Data Entry",width=15)
cancel_check.grid(row=1,column=0,columnspan=2, padx=2,pady=2,sticky="nsew")
cancel_options=StringVar()
cancel_code=ttk.Combobox(lf4,textvariable=cancel_options,state='raedonly')
cancel_code.grid(row=2,column=0,padx=10,pady=2,sticky="nsew")
cancel_code['values']=("Clear Day & Slot ","Clear All Data","Clear Data from Time Table")
cancel_code.current()
cancel_code.bind("<<ComboboxSelected>>",cancel_option_selection)

cancel_button=Button(lf4,text="Confirm CANCEL",bg="white",command=confirm_cancel)
cancel_button.grid(row=2,column=1,padx=2,pady=2,sticky="nsew")
verify_data=Label(lf4,text="Verify & Update TimeTable",width=15)
verify_data.grid(row=1,column=2,columnspan=2, padx=2,pady=2,sticky="nsew")
verify_button=Button(lf4,text="Verified OK",bg="white",command=verify_ok)
verify_button.grid(row=2,column=2,padx=2,pady=2,sticky="nsew")
update_button=Button(lf4,text="Update TimeTable",bg="white",command=update_tt)
update_button.grid(row=2,column=3,padx=2,pady=2,sticky="nsew")
root.mainloop()
#