# Timetable project
# Intermediate Stage
from tkinter import*
from tkinter import ttk
import tkinter as tk
from openpyxl import load_workbook,Workbook
import pandas as pd
import numpy as np
from tt_show_class_jan20 import TimeTable
import os
def create_excelsheet_orread(file_name,sheet_name):
    if os.path.exists(file_name):
        sheet_names_list = pd.ExcelFile(file_name).sheet_names
        if sheet_name not in sheet_names_list:
            df_init = pd.DataFrame(data_ref)
            with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                df_init.to_excel(writer,sheet_name=sheet_name,index=False,engine = "openpyxl")
            df = pd.read_excel(file_name,sheet_name=sheet_name,engine="openpyxl")
            return(df)
        else: 
            df = pd.read_excel(file_name,sheet_name=sheet_name,engine="openpyxl")
            return(df)
    else:
        df_init = pd.DataFrame(data_ref)
        df_init.to_excel(file_name,sheet_name,index=False,engine = "openpyxl")
        df = pd.read_excel(file_name,sheet_name=sheet_name,engine="openpyxl")
        return(df)   
def find_key(dictionary,value):
    for key,val in dictionary.items():
        if val == value:
            return(key)
    return None
# Combining sublist in data frame and eliminating common ref elements from two lists
def common_lst(df_ref,lstname_ref,lstname_sub):  
    comn_ref_lst = []
    comn_sub_lst = []
    for i in range(len(df_ref)):
        print(len(df_ref.at[df_ref.index.tolist()[i],lstname_ref]))
        if len(df_ref.at[df_ref.index.tolist()[i],lstname_ref].split(",")) > 1:
            ref_lst = df_ref.at[df_ref.index.tolist()[i],lstname_ref].split(",")
            sub_lst = df_ref.at[df_ref.index.tolist()[i],lstname_sub].split(",")
        else:
            ref_lst = [str(df_ref.at[df_ref.index.tolist()[i],lstname_ref])]
            sub_lst = [str(df_ref.at[df_ref.index.tolist()[i],lstname_sub])]
        comn_ref_lst += ref_lst
        comn_sub_lst += sub_lst                   
    unifrom_ref_list = []
    unifrom_sub_list = []
    for i in range(len(comn_ref_lst)):
        if comn_ref_lst[i] not in unifrom_ref_list:
            unifrom_ref_list.append(comn_ref_lst[i])
            unifrom_sub_list.append(comn_sub_lst[i])                 
    return(unifrom_sub_list)
    """
def excel_read1(file_name,sheet_name):
    data_dfrow = []
    lst_dfcol = []
    workbook = load_workbook(filename=f"{file_name}.xlsx")
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rowdf_dict = dict(zip(headers, row))
        data_dfrow.append(rowdf_dict)
    for column in sheet.iter_cols(min_row=2, values_only=True):
        df_col = []
        for cell in column:
            df_col.append(cell)
        lst_dfcol.append(df_col)
        coldf_dict = dict(zip(headers, lst_dfcol))
    workbook.close()
    row_df_ret = [data_dfrow,coldf_dict,headers]
    return row_df_ret
def excel_writerows(file_name,sheet_name,data_header,data_column):
    workbook = load_workbook(filename=f"{file_name}.xlsx")
    sheet = workbook[sheet_name]
    print(sheet.title)
    return
"""
def parent_dept_selection(event):
    global df_dept
    action_code.set('')
    action_code.config(values=(''))
    dept_emp_code.set('')
    df_dept = create_excelsheet_orread(file_name='faculty_data.xlsx',sheet_name=dept_code.get())
    dept_emp_code.config(values=df_dept.emp_code.tolist())
def emp_code_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    ltp_code.set('')
    ltp_code.config(values=(''))
    assign_dept_code.set('')
    assign_dept_code.config(values=("CE","ME","EE"))
def dept_option_selection(event):
    odd_even_code.config(values=("ODD","EVEN"))
    action_code.set('')
    action_code.config(values=('')) 
def odd_even_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    odd_even = odd_even_code.get()
    if odd_even == 'ODD':
        odd_even_list = ['s1','s3','s5','s7']
    elif odd_even == 'EVEN':
        odd_even_list = ['s2','s4','s6','s8']
    else:
        pass
    semester_code.set('')
    semester_code.config(values=odd_even_list)
def semester_option_selection(event):
    action_code.set('')
    action_code.config(values=(''))
    ltp_code.set('')
    ltp_code.config(values=("Lecture","Tutorial","Practical","Projects"))
def ltp_option_selection(event):
    global faculty_name,df_sel,df_curi_active,code_lst_lab,code_lst_tut
    action_code.set('')
    action_code.config(values=(''))
    """
    global data_faculty,data_tt,data_temp,faculty_sel
    global df_dept,df_curi_lec,df_curi_tut
    data_curi = excel_read1(f"curiculam_{assign_dept_code.get()}",semester_code.get())
    new_row_lst = [data_curi[0][i] for i in range(len(data_curi[1]["Select"])) if data_curi[1]["Select"][i] != 0]
    data_col_lst = list(data_curi[1].values())  
    cur_col_lst = [[data_col_lst[j][i] for i in range(len(data_curi[1]["Select"])) if data_curi[1]["Select"][i] != 0] for j in range(len(data_curi[2]))]
    cur_col_dict = dict(zip(data_curi[2], cur_col_lst))
    data_tt = excel_read1(f"timeTable_{assign_dept_code.get()}",semester_code.get())
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    """
    faculty_name = df_dept.nick_name.tolist()[df_dept.emp_code.tolist().index(int(dept_emp_code.get()))]
    df_curi = create_excelsheet_orread(f"curiculam_{assign_dept_code.get()}.xlsx",semester_code.get())
    df_curi_active = df_curi[df_curi["Select"] != 0]
    df_sel = create_excelsheet_orread(f"faculty_assignment_{odd_even_code.get()}.xlsx",sheet_name=assign_dept_code.get())
    match ltp_code.get():
        case "Lecture":
            df_curi_lec = df_curi_active[df_curi_active["L"] != 0] 
            df_sel_lec = df_sel[["nick_name","deptL_class","semL","L_code"]].copy()         
            print(df_sel_lec)
            if df_sel_lec.empty:
                code_lst = df_curi_lec.Code.tolist()
            else:
                df_sel_lec = df_sel[["nick_name","deptL_class","semL","L_code"]].copy()
                df_sel_lec.dropna(subset=['nick_name'],inplace=True) # remove rows with empty nick_name
                if faculty_name in df_sel_lec.nick_name.tolist():
                    df_sel_lec=df_sel_lec[df_sel_lec["nick_name"]!=faculty_name]
                    code_lst_temp=df_curi_lec.Code.tolist()
                    for i in range(len(common_lst(df_sel_lec,"L_code","L_code"))):    
                        if assign_dept_code.get() == common_lst(df_sel_lec,"L_code","deptL_class")[i] and semester_code.get() == common_lst(df_sel_lec,"L_code","semL")[i]:
                            code_lst_temp.remove(common_lst(df_sel_lec,"L_code","L_code")[i])
                else:
                    code_lst_temp=df_curi_lec.Code.tolist()
                    for i in range(len(common_lst(df_sel_lec,"L_code","L_code"))): 
                        if assign_dept_code.get() == common_lst(df_sel_lec,"L_code","deptL_class")[i] and semester_code.get() == common_lst(df_sel_lec,"L_code","semL")[i]:
                            code_lst_temp.remove(common_lst(df_sel_lec,"L_code","L_code")[i])
                df_sel_mylec = df_sel[["nick_name","L_code","L_class","L_hr"]].copy()
                df_sel_mylec.dropna(subset=['nick_name'],inplace=True) # remove rows with empty nick_name
                if faculty_name in df_sel_mylec.nick_name.tolist():
                    df_sel_mylec=df_sel_mylec[df_sel_mylec["nick_name"]==faculty_name]
                    for i in range(len(common_lst(df_sel_mylec,"L_code","L_code"))):
                        if common_lst(df_sel_mylec,"L_code","L_class")[i] == common_lst(df_sel_mylec,"L_code","L_hr")[i]:
                            code_lst_temp.remove(common_lst(df_sel_mylec,"L_code","L_code")[i])
                code_lst=code_lst_temp    
                print(code_lst)
        case "Tutorial":
            df_curi_tut = df_curi_active[df_curi_active["T"] != 0]
            df_sel_tut = df_sel[["nick_name","deptT_class","semT","T_code","T_hr"]].copy()
            df_sel_tut.dropna(subset=['T_code'],inplace=True) # remove rows with empty tutorial code
            if df_sel_tut.empty:
                code_lst = df_curi_tut.Code.tolist()
                code_lst_tut = code_lst
            else:
                T_count = []
                for i in range(len(df_curi_tut)):
                    if list(df_curi_tut.Number)[i] <= 26:
                        T_need = 1
                    elif list(df_curi_tut.Number)[i] <= 48:
                        T_need = 2
                    else:
                        T_need = 3
                    T_count.append(T_need)
                code_lst_tut = df_curi_tut.Code.tolist()
                for i in range(len(df_curi_tut)):
                    T_used = 0
                    for j in range(len(df_sel_tut)):
                        if df_curi_tut.Code.tolist()[i] in df_sel_tut.at[df_sel_tut.index.tolist()[j],"T_code"]:  
                            if assign_dept_code.get() == common_lst(df_sel_tut,"T_code","deptT_class")[j] and semester_code.get() == common_lst(df_sel_tut,"T_code","semT")[j]:  
                                T_used += 1
                            if T_used == int(T_count[df_curi_tut.Code.tolist().index(df_curi_tut.Code.tolist()[i])]):                        
                                code_lst_tut.remove(df_curi_tut.Code.tolist()[i])
                df_sel_mytut = df_sel_tut[df_sel_tut["nick_name"]==faculty_name]
                code_lst = code_lst_tut
                for i in range(len(common_lst(df_sel_mytut,"T_code","T_code"))):
                    print(common_lst(df_sel_mytut,"T_code","T_code")[i])
                    if common_lst(df_sel_mytut,"T_code","T_code")[i] in code_lst_temp:
                        code_lst.remove(common_lst(df_sel_mytut,"T_code","T_code")[i])
                        code_lst_tut = code_lst
        case "Practical":
            df_curi_lab = df_curi_active[df_curi_active["P"] != 0]
            df_sel_lab = df_sel[["nick_name","deptP_class","semP","P_code","P_hr"]].copy()
            df_sel_lab.dropna(subset=['P_code'],inplace=True) # remove rows with empty tutorial code
            if df_sel_lab.empty:
                code_lst = df_curi_lab.Code.tolist()
                code_lst_lab = code_lst        
            else:
                P_count = []
                for i in range(len(df_curi_lab)):
                    if list(df_curi_lab.Number)[i] <= 18:
                        P_need = 1
                    elif list(df_curi_lab.Number)[i] <= 32:
                        P_need = 2
                    elif list(df_curi_lab.Number)[i] <= 48:
                        P_need = 3
                    else:
                        P_need = 4                
                    P_count.append[P_need]          
                code_lst_lab = df_curi_lab.Code.tolist()
                for i in range(len(df_curi_lab)):
                    P_used = 0
                    for j in range(len(df_sel_lab)):
                        if df_curi_lab.Code.tolist()[i] in df_sel_lab.at[df_sel_lab.index.tolist()[j],"P_code"]:  
                            if assign_dept_code.get() == common_lst(df_sel_lab,"P_code","deptP_class")[j] and semester_code.get() == common_lst(df_sel_tut,"P_code","semP")[j]:  
                                P_used += 1
                            if P_used == int(P_count[df_curi_lab.Code.tolist().index(df_curi_lab.Code.tolist()[i])]):                        
                                code_lst_lab.remove(df_curi_lab.Code.tolist()[i])
                df_sel_mylab = df_sel_lab[df_sel_lab["nick_name"]==faculty_name]
                code_lst = code_lst_lab                      
                for i in range(len(common_lst(df_sel_mylab,"P_code","P_code"))):
                    print(common_lst(df_sel_mylab,"P_code","P_code")[i])
                    if common_lst(df_sel_mylab,"P_code","P_code")[i] in code_lst_lab:
                        code_lst.remove(common_lst(df_sel_mylab,"P_code","P_code")[i])             
        case "Projects":
            code_lst = ["Guide-4","Guide-8","Other HoD","other CT","Special"]
            faculty_in_guide_sel = df_sel[df_sel['nick_name'] == faculty_name]
            if list(faculty_in_guide_sel.R_code) != []:
                code_lst_l0 = code_lst
                code_lst = code_lst_l0
            print(code_lst)
            for i in range(len(code_lst)):
                R_ref = [round(int(list(df_curi_active.Number)[i])/4),round(int(list(df_curi_active.Number)[i])/8),1,1,2]
            print(R_ref)
        case _:
            print("Who")
    sub_code.set('')
    sub_code.config(values=code_lst)
def sub_option_selection(event):
    action_code.config(values=("PROCEED","CLEAR"))
def action_option_selection(event):
    global data_tt,data_faculty
    global class_x,data,ro_co_dict,row_val
    action_entry=action_code.get()
    if action_entry == "CLEAR":
        lf1_lf2_clear()
        pass
    elif action_entry == "PROCEED":
        TimeTable(root,f"timeTable_{assign_dept_code.get()}.xlsx",semester_code.get())
        ro_co_dict = {}
        for i in range(5):
            ro_co_dict.update({i:[0 for _ in range(7)]})
        print(ro_co_dict)
        print(table_tt)
        week_click(None)
        refresh_button.config(bg="Blue",fg="white")
        text_box2.delete('1.0','end')
        text_box2.insert('1.0',f"Select weekday and one from available slot before REFRESH")
def week_click(week_value:any):
    global code_lst_lab,code_lst_tut,faculty_name
    df_tt = pd.read_excel(f"timeTable_{assign_dept_code.get()}.xlsx",sheet_name=semester_code.get(),engine="openpyxl")
    cb1.config(state='disabled')
    cb2.config(state='disabled')
    cb3.config(state='disabled')
    cb4.config(state='disabled')
    cb5.config(state='disabled')
    cb6.config(state='disabled')
    cb7.config(state='disabled')
    sr0.set(0)
    sr1.set(0)
    sr2.set(0)
    sr3.set(0)
    sr4.set(0)
    sr5.set(0)
    sr6.set(0)
    match ltp_code.get():
        case "Lecture":
            if code_lst_lab == []:
                if df_tt.at[r3.get(),"slot_1"] == "FREE":
                    cb1.config(state='normal')
                if df_tt.at[r3.get(),"slot_2"] == "FREE":
                    cb2.config(state='normal')
                if df_tt.at[r3.get(),"slot_3"] == "FREE":
                    cb3.config(state='normal')
                if df_tt.at[r3.get(),"slot_4"] == "FREE":
                    cb4.config(state='normal')
                if df_tt.at[r3.get(),"slot_5"] == "FREE":
                    cb5.config(state='normal')
                if df_tt.at[r3.get(),"slot_6"] == "FREE":
                    cb6.config(state='normal')
                if df_tt.at[r3.get(),"slot_x"] == "FREE":
                    cb6.config(state='normal')
            else:
                print("NOT blank")
                text_box2.delete('1.0','end')
                text_box2.insert('1.0',f"!! WARNING !! First complete the entry in LAB slots to get all slots for the LECTURE")
                if df_tt.at[r3.get(),"slot_1"] == "FREE":
                    cb1.config(state='normal')
                if df_tt.at[r3.get(),"slot_2"] == "FREE":
                    cb2.config(state='normal')           
        case "Tutorial":
            df_sel_lec = df_sel[["nick_name","deptL_class","semL","L_code"]].copy()
            df_sel_mylec = df_sel_lec[df_sel_mylec['nick_name'] == faculty_name]
            if sub_code.get() in common_lst(df_sel_mylec,"L_code","L_code"):
                for i in range(len(common_lst(df_sel_lec,"L_code","L_code"))):    
                    if assign_dept_code.get() == common_lst(df_sel_lec,"L_code","deptL_class")[i] and semester_code.get() == common_lst(df_sel_lec,"L_code","semL")[i]:
                        if code_lst_lab == []:
                            if df_tt.at[r3.get(),"slot_1"] == "FREE":
                                cb1.config(state='normal')
                            if df_tt.at[r3.get(),"slot_2"] == "FREE":
                                cb2.config(state='normal')
                            if df_tt.at[r3.get(),"slot_3"] == "FREE":
                                cb3.config(state='normal')
                            if df_tt.at[r3.get(),"slot_4"] == "FREE":
                                cb4.config(state='normal')
                            if df_tt.at[r3.get(),"slot_5"] == "FREE":
                                cb5.config(state='normal')
                            if df_tt.at[r3.get(),"slot_6"] == "FREE":
                                cb6.config(state='normal')
                            if df_tt.at[r3.get(),"slot_x"] == "FREE":
                                cb6.config(state='normal')
                        else:
                            text_box2.delete('1.0','end')
                            text_box2.insert('1.0',f"!! WARNING !! First complete the entry in LAB slots to get all slots for the LECTURE")
                            if df_tt.at[r3.get(),"slot_1"] == "FREE":
                                cb1.config(state='normal')
                            if df_tt.at[r3.get(),"slot_2"] == "FREE":
                                cb2.config(state='normal')   
                else:
                    row_tut, col_tut = np.where(df_tt == sub_code.get())
                    text_box2.delete('1.0','end')
                    text_box2.insert('1.0',f"!! ENTER !! The choice in Timetable {row_tut}, {col_tut}")
        case "Practical":
            df_sel_lec = df_sel[["nick_name","deptL_class","semL","L_code"]].copy()
            df_sel_mylec = df_sel_lec[df_sel_mylec['nick_name'] == faculty_name]
            if sub_code.get() in common_lst(df_sel_mylec,"L_code","L_code"):
                for i in range(len(common_lst(df_sel_lec,"L_code","L_code"))):    
                    if assign_dept_code.get() == common_lst(df_sel_lec,"L_code","deptL_class")[i] and semester_code.get() == common_lst(df_sel_lec,"L_code","semL")[i]:
                        if df_tt.at[r3.get(),"slot_3"] == "FREE":
                            cb3.config(state='normal')
                        if df_tt.at[r3.get(),"slot_4"] == "FREE":
                            cb3.config(state='normal')
                        if df_tt.at[r3.get(),"slot_5"] == "FREE":
                            cb5.config(state='normal')
                        if df_tt.at[r3.get(),"slot_6"] == "FREE":
                            cb5.config(state='normal')
                        if df_tt.at[r3.get(),"slot_x"] == "FREE":
                            cb6.config(state='normal')
                else:
                    row_tut, col_tut = np.where(df_tt == sub_code.get())
                    text_box2.delete('1.0','end')
                    text_box2.insert('1.0',f"!! ENTER !! The choice in Timetable {row_tut}, {col_tut}")
    verify_button.config(bg="white",fg="black")
def refresh_weekslot():
    ro_co_dict[r3.get()] = [sr0.get(),sr1.get(),sr2.get(),sr3.get(),sr4.get(),sr5.get(),sr6.get()]
    if 1 not in ro_co_dict[r3.get()]:
        text_box2.insert('1.0',f"!!WARNING!! Select weekday and one from available slot before REFRESH...................")
        return
    #print(r3.get(),ro_co_dict[r3.get()].index(1))
    refresh_button.config(bg="yellow",fg="black")
    verify_button.config(bg="Blue",fg="white")
    update_button.config(bg="white",fg="black")
    text_box2.delete('1.0','end')
    text_box2.insert('1.0',f"Confirm your choice {ltp_code.get()} : for the subject {sub_code.get()} for {semester_code.get()} in {assign_dept_code.get()} .... ")
    text_box2.insert('2.0',f"Selected choice {data_tt[1]["week_day"][r3.get()]} for the slot {data_tt[2][(ro_co_dict[r3.get()].index(1))+1]} ")
def verify_ok():
    update_button.config(bg="Blue",fg="white")
    verify_button.config(bg="white",fg="black")
    refresh_button.config(bg="white",fg="black")
    text_box2.delete('1.0','end')
    text_box2.insert('1.0',f"Confirm and SAVE to Time Table")
def update_tt():
    global faculty_name,df_sel,df_curi_active
    update_button.config(bg="white",fg="black")
    match ltp_code.get():
        case "Lecture":
            df_curi_lec = df_curi_active[df_curi_active["L"] != 0]
            class_hr_index = df_curi_lec.Code.tolist().index(sub_code.get())
            lec_class_hr = str(df_curi_lec.L.tolist()[class_hr_index])
            df_sel_lec = df_sel[["nick_name","deptL_class","semL","L_code","L_class","L_hr"]].copy()
            if faculty_name in df_sel_lec.nick_name.tolist():
                df_sel_mylec = df_sel_lec[df_sel_lec["nick_name"]==faculty_name]
                index_row = df_sel_lec.nick_name.tolist().index(faculty_name)
                if sub_code.get() in common_lst(df_sel_lec,"L_code","L_code"):                           #df_sel_mylec,??????????????????????????
                    index_col = common_lst(df_sel_mylec,"L_code","L_code").index(sub_code.get())
                    if assign_dept_code.get()==common_lst(df_sel_mylec,"L_code","deptL_class")[index_col] and common_lst(df_sel_mylec,"L_code","semL")[index_col]:   
                        new_hr_lst = []
                        for i in range(len(common_lst(df_sel_mylec,"L_code","L_hr"))):
                            if i == index_col:
                                new_hr_lst.append(int(common_lst(df_sel_mylec,"L_code","L_hr")[i])+1)
                            else:
                                new_hr_lst.append(int(common_lst(df_sel_mylec,"L_code","L_hr")[i]))                              
                        df_sel.loc[index_row,"L_hr"] = ",".join(str(x) for x in new_hr_lst)
                        df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
                else:
                    df_sel.loc[index_row,"deptL_class"] = f"{df_sel_lec.at[index_row,"deptL_class"]},{assign_dept_code.get()}"
                    df_sel.loc[index_row,"semL"] = f"{df_sel_lec.at[index_row,"semL"]},{semester_code.get()}"
                    df_sel.loc[index_row,"L_code"] = f"{df_sel_lec.at[index_row,"L_code"]},{sub_code.get()}"
                    df_sel.loc[index_row,"L_class"] = f"{df_sel_lec.at[index_row,"L_class"]},{lec_class_hr}"
                    df_sel.loc[index_row,"L_hr"] = f"{df_sel_lec.at[index_row,"L_hr"]},{1}"
                    df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")                   
            else:
                new_record = pd.DataFrame([{'nick_name':faculty_name,'emp_code':dept_emp_code.get(),'dept_origin':dept_code.get(),
                    'deptL_class':assign_dept_code.get(),'semL':semester_code.get(),'L_code':sub_code.get(),'L_class':lec_class_hr,"L_hr":1}])
                df_sel = pd.concat([df_sel,new_record],ignore_index=True)    
                df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
        case "Tutorial":
            #df_curi_tut = df_curi_active[df_curi_active["T"] != 0]
            df_sel_tut = df_sel[["nick_name","deptT_class","semT","T_code","T_hr"]].copy()
            if faculty_name in list(df_sel_tut.nick_name):
                df_sel_mytut = df_sel_tut[df_sel_tut["nick_name"]==faculty_name]
                index_row = df_sel_tut.nick_name.tolist().index(faculty_name)
                if df_sel_mytut.T_code.any():
                    """
                    print("yes data")
                    if sub_code.get() in common_lst(df_sel_tut,"T_code","T_code"):
                        if semester_code.get() not in df_sel_tut.at[index_new,"semT"].split(","):
                            if assign_dept_code.get() not in df_sel_tut.at[index_new,"deptT_class"].split(","):
                                """
                    df_sel.loc[index_row,"deptT_class"] = f"{df_sel_tut.at[index_row,"deptT_class"]},{assign_dept_code.get()}"
                    df_sel.loc[index_row,"semT"] = f"{df_sel_tut.at[index_row,"semT"]},{semester_code.get()}"
                    df_sel.loc[index_row,"T_code"] = f"{df_sel_tut.at[index_row,"T_code"]},{sub_code.get()}"
                    df_sel.loc[index_row,"T_hr"] = f"{df_sel_tut.at[index_row,"T_hr"]},{1}"
                    df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
                else:
                    df_sel.loc[index_row,"deptT_class"] = assign_dept_code.get()
                    df_sel.loc[index_row,"semT"] = semester_code.get()
                    df_sel.loc[index_row,"T_code"] = sub_code.get()
                    df_sel.loc[index_row,"T_hr"] =f"{1}"
                    df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
            else:
                new_record = pd.DataFrame([{'nick_name':faculty_name,'emp_code':dept_emp_code.get(),'dept_origin':dept_code.get(),
                    'deptT_class':assign_dept_code.get(),'semT':semester_code.get(),'T_code':sub_code.get(),"T_hr":1}])
                df_sel = pd.concat([df_sel,new_record],ignore_index=True)    
                df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")               
        case "Practical":
            df_sel_lab = df_sel[["nick_name","deptP_class","semP","P_code","P_hr"]].copy()
            if faculty_name in list(df_sel_lab.nick_name):
                df_sel_mylab = df_sel_lab[df_sel_lab["nick_name"]==faculty_name]
                index_row = df_sel_lab.nick_name.tolist().index(faculty_name)
                print(df_sel_mylab.P_code.any())
                if df_sel_mylab.P_code.any():
                    df_sel.loc[index_row,"deptP_class"] = f"{df_sel_lab.at[index_row,"deptP_class"]},{assign_dept_code.get()}"
                    df_sel.loc[index_row,"semP"] = f"{df_sel_lab.at[index_row,"semP"]},{semester_code.get()}"
                    df_sel.loc[index_row,"P_code"] = f"{df_sel_lab.at[index_row,"P_code"]},{sub_code.get()}"
                    df_sel.loc[index_row,"P_hr"] = f"{df_sel_lab.at[index_row,"P_hr"]},{1}"
                    df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
                else:
                    df_sel.loc[index_row,"deptP_class"] = assign_dept_code.get()
                    df_sel.loc[index_row,"semP"] = semester_code.get()
                    df_sel.loc[index_row,"P_code"] = sub_code.get()
                    df_sel.loc[index_row,"P_hr"] =f"{1}"
                    df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl")
            else:
                new_record = pd.DataFrame([{'nick_name':faculty_name,'emp_code':dept_emp_code.get(),'dept_origin':dept_code.get(),
                    'deptP_class':assign_dept_code.get(),'semP':semester_code.get(),'P_code':sub_code.get(),"P_hr":1}])
                df_sel = pd.concat([df_sel,new_record],ignore_index=True)    
                df_sel.to_excel(f"faculty_assignment_{odd_even_code.get()}.xlsx",dept_code.get(),index=False,engine="openpyxl") 
    dept_emp_code.set('')
    assign_dept_code.set('')
    semester_code.set('')
    ltp_code.set('')
    sub_code.set('')
    action_code.set('')                   
def cancel_option_selection(event):
    lf1_lf2_clear()
    pass
def lf1_lf2_clear():
    dept_emp_code.set('')
    dept_emp_code['values']=()
    assign_dept_code.set('')
    odd_even_code.set('')
    semester_code.set('')
    ltp_code.set('')
    sub_code.set('')
    cancel_code.set('')

def common_for_cancel_options():
    pass

def clear_day_slot():
    text_box1.delete('1.0','end')
    text_box2.delete('1.0','end')
    sr0.set(0)
    sr1.set(0)
    sr2.set(0)
    sr3.set(0)
    sr4.set(0)
    sr5.set(0)
    sr6.set(0)

def cancel_option(event):
    global clear_status
    cancel_entry = cancel_click.get()
    if cancel_entry == "Assigned Class":
        clear_status = 1
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")    
    elif cancel_entry == "Day & Slot":
        clear_status = 2
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")
    else:  
        clear_status = 3
        common_for_cancel_options()
        text_box1.insert('1.0',"All the data entry under 1)Assigned for Dept./Topic 2)semester 3)Class Type 4)Subject Code will be reset for fresh entry.")
def confirm_cancel():
    pass

root=Tk()
root.title("Data entry by Faculty for TimeTable")
w_width = root.winfo_screenwidth()
w_height = root.winfo_screenheight()
dev_WIDTH = 700
dev_HEIGHT = 550
ref_s_WIDTH = 1920
ref_s_HEIGHT = 1080
s_width = int(dev_WIDTH*w_width/ref_s_WIDTH)
s_height = int(dev_HEIGHT*w_height/ref_s_HEIGHT)
#print(s_width,s_height)
root.geometry("700x550")
root.iconbitmap('Logo.ico')
root.columnconfigure(0,weight=1) #weight= 1 indicates scale of one
root.rowconfigure(0,weight=1)
frame=Frame(root)
frame.grid(row=0,column=0,padx=10,pady=5,sticky="nsew")
x = "Lecture"
global table_tt
table_tt=[{"week_day":" ","slot_1":" ","slot_2":" ","slot_3":" ",
    "slot_4":" ","slot_5":" ","slot_6":" ","slot_x":" "}]
global code_lst_lab
code_lst_lab = []
odd_even="odd"
day_slot={"week_slot":3,"time_slot0":0,"time_slot1":1,"time_slot2":2,"time_slot3":3,
    "time_slot4":4,"time_slot5":5,"time_slot6":6}
sr=IntVar()
#Storage file for manipulation
data_ref = [{'nick_name':'','emp_code':'','dept_origin':'','deptL_class':'','semL':'','L_code':'','L_class':'',"L_hr":'',
    'deptT_class':'','semT':'','T_code':'',"T_hr":'','deptP_class':'','semP':'','P_code':'',"P_hr":'',
    'deptR_class':'','semR':'','R_code':'',"R_hr":''}]
#Storage file on departments
dept_data = [{'emp_code':'','dept':'','grade':'','nick_name':'','pay_band':''}]
#Storage file on curiculam
curi_data = [{'Code':'','Subject':'','Select':'','Number':'','L':'','T':'','P':'','R':''}]

#LBELFRAME lf1
frame.columnconfigure(0,weight=1) #weight= 1 indicates scale of one
frame.rowconfigure((0,1,2,3),weight=1)
lf1=LabelFrame(frame,text="Employ Code: Parent Dept: Faculty Initial: Assigned Dept.",padx=5,pady=2,fg="Blue")
lf1.grid(row=0,column=0,padx=10,pady=5,sticky="nsew")
lf1.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf1.rowconfigure((0,1),weight=1)
dept_emp_code=Label(lf1,text="Parent Dept.& Emp.Code",width=15)
dept_emp_code.grid(row=0,column=0,columnspan=2,padx=2,pady=2,sticky="nsew")
dept_class=Label(lf1,text="Assigned Department/Semester",width=15)
dept_class.grid(row=0,column=2,columnspan=2,padx=2,pady=2,sticky="nsew")
sub_lec_faculty = {}
sub_tut_faculty = {}
sub_lab_faculty = {}
sub_proj_faculty = {}
#lf1 entries
dept_options=StringVar()
dept_code=ttk.Combobox(lf1,textvariable=dept_options,state='raedonly')
dept_code.grid(row=1,column=0,padx=10,pady=2,sticky="nsew")
dept_code['values']=("CE","ME","EE")
dept_code.current()
dept_code.bind("<<ComboboxSelected>>",parent_dept_selection)
dept_emp_options=StringVar()
dept_emp_code=ttk.Combobox(lf1,textvariable=dept_emp_options,state='raedonly')
dept_emp_code.grid(row=1,column=1,padx=10,pady=2,sticky="nsew")
dept_emp_code['values']=()
dept_emp_code.current()
dept_emp_code.bind("<<ComboboxSelected>>",emp_code_selection)
assign_dept_options=StringVar()
assign_dept_code=ttk.Combobox(lf1,textvariable=assign_dept_options,state='raedonly')
assign_dept_code.grid(row=1,column=2,padx=10,pady=2,sticky="nsew")
assign_dept_code['values']=()
assign_dept_code.current()
assign_dept_code.bind("<<ComboboxSelected>>",dept_option_selection)
odd_even_options=StringVar()
odd_even_code=ttk.Combobox(lf1,textvariable=odd_even_options,state='raedonly')
odd_even_code.grid(row=1,column=3,padx=10,pady=2,sticky="nsew")
odd_even_code['values']=()
odd_even_code.current()
odd_even_code.bind("<<ComboboxSelected>>",odd_even_selection)
#LBELFRAME lf2
lf2=LabelFrame(frame,text="Selecting Assigned Classes",padx=5,pady=2,fg="Blue")
lf2.grid(row=1,column=0,padx=10,pady=5,sticky="nsew")
#lf2 options
global semester_options
#lf2 labels
lf2.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf2.rowconfigure((0,1,2),weight=1)
text_box1=Text(lf2, width=60,height=3,fg="red",font=("Ariel",11))
text_box1.grid(row=0, column=0,columnspan=4,padx=10,pady=5,sticky="nsew")
semester_class=Label(lf2,text="Semester",width=15)
semester_class.grid(row=1,column=0,padx=2,pady=2,sticky="nsew")
ltp_type=Label(lf2,text="Class type",width=15)
ltp_type.grid(row=1,column=1,padx=2,pady=2,sticky="nsew")
sub_code=Label(lf2,text="Subject Code",width=15)
sub_code.grid(row=1,column=2,padx=2,pady=2,sticky="nsew")
sub_check=Label(lf2,text="ACTION",width=15)
sub_check.grid(row=1,column=3,padx=2,pady=2,sticky="nsew")
#lf2 entries
semester_options=StringVar()
semester_code=ttk.Combobox(lf2,textvariable=semester_options,state='raedonly')
semester_code.grid(row=2,column=0,padx=10,pady=2,sticky="nsew")
semester_code['values']=()
semester_code.current()
semester_code.bind("<<ComboboxSelected>>",semester_option_selection)
ltp_options=StringVar()
ltp_code=ttk.Combobox(lf2,textvariable=ltp_options,state='raedonly')
ltp_code.grid(row=2,column=1,padx=10,pady=2,sticky="nsew")
ltp_code['values']=()
ltp_code.current()
ltp_code.bind("<<ComboboxSelected>>",ltp_option_selection)
sub_options=StringVar()
sub_code=ttk.Combobox(lf2,textvariable=sub_options,state='raedonly')
sub_code.grid(row=2,column=2,padx=10,pady=2,sticky="nsew")
sub_code['values']=()
sub_code.current()
sub_code.bind("<<ComboboxSelected>>",sub_option_selection)
action_options=StringVar()
action_code=ttk.Combobox(lf2,textvariable=action_options,state='raedonly')
action_code.grid(row=2,column=3,padx=10,pady=2,sticky="nsew")
action_code['values']=("CLEAR")
action_code.current()
action_code.bind("<<ComboboxSelected>>",action_option_selection)
#LBELFRAME lf3
lf3=LabelFrame(frame,text="Timetable Weekday and Slot selection",padx=5,pady=2,fg="Blue")
lf3.grid(row=2,column=0,padx=10,pady=5,sticky="nsew")
lf3.columnconfigure((0,1,2,3,4,5,6),weight=1,uniform="a") #weight= 1 indicates scale of one
lf3.rowconfigure((0,1,2),weight=1)
r3=IntVar()
week_day=["Monday","Tuesday","Wednesday","Thursday","Friday"]
for i in range(len(week_day)):
    week=(Radiobutton(lf3,text=week_day[i],variable=r3,value=i,command=lambda:week_click(r3.get())))
    week.grid(row=0,column=i+1,padx=10,pady=2,sticky="nsew")
refresh_button = Button(lf3,text="REFRESH",command=refresh_weekslot)
refresh_button.grid(row=0,column=6,padx=2,pady=2,sticky="nsew")
#print(week.r3)
show_msg=Label(lf3,text="-----------Slot 1 to Slot 6 for Lecture & Tutorial and Slot X optional-------------",fg="Blue")
show_msg.grid(row=1, column=0, columnspan=7,padx=2,pady=2,sticky="nsew")
slots=["Slot 1","Slot 2","Slot 3","Slot 4","Slot 5","Slot 6","Slot X"]
sr0 = IntVar()
sr1 = IntVar()
sr2 = IntVar()
sr3 = IntVar()
sr4 = IntVar()
sr5 = IntVar()
sr6 = IntVar()
global cb1
cb1=Checkbutton(lf3, text=slots[0], variable=sr0,state='disabled')
cb1.grid(row=3, column=0,sticky="nsew")
cb2=Checkbutton(lf3, text=slots[1], variable=sr1,state='disabled')
cb2.grid(row=3, column=1,sticky="nsew")
cb3=Checkbutton(lf3, text=slots[2], variable=sr2,state='disabled')
cb3.grid(row=3, column=2,sticky="nsew")
cb4=Checkbutton(lf3, text=slots[3], variable=sr3,state='disabled')
cb4.grid(row=3, column=3,sticky="nsew")
cb5=Checkbutton(lf3, text=slots[4], variable=sr4,state='disabled')
cb5.grid(row=3, column=4,sticky="nsew")
cb6=Checkbutton(lf3, text=slots[5], variable=sr5,state='disabled')
cb6.grid(row=3, column=5,sticky="nsew")
cb7=Checkbutton(lf3, text=slots[6], variable=sr6,state='disabled')
cb7.grid(row=3, column=6,sticky="nsew")
#LBELFRAME lf4
#lf4 options
lf4=LabelFrame(frame,text="Timetable Data Entry Commands",padx=5,pady=2,fg="Blue")
lf4.grid(row=3,column=0,padx=10,pady=5,sticky="nsew")
lf4.columnconfigure((0,1,2,3),weight=1,uniform="a") #weight= 1 indicates scale of one
lf4.rowconfigure((0,1),weight=1)
text_box2=Text(lf4, width=60,height=3,fg="red",font=("Ariel",11))
text_box2.grid(row=0,column=0,columnspan=4,padx=10,pady=5,sticky="nsew")
cancel_check=Label(lf4,text="Cancel Data Entry",width=15)
cancel_check.grid(row=1,column=0,columnspan=2, padx=2,pady=2,sticky="nsew")
cancel_options=StringVar()
cancel_code=ttk.Combobox(lf4,textvariable=cancel_options,state='raedonly')
cancel_code.grid(row=2,column=0,padx=10,pady=2,sticky="nsew")
cancel_code['values']=("Clear Day & Slot ","Clear All Data","Clear Data from Time Table")
cancel_code.current()
cancel_code.bind("<<ComboboxSelected>>",cancel_option_selection)

cancel_button=Button(lf4,text="Confirm CANCEL",bg="white",command=confirm_cancel)
cancel_button.grid(row=2,column=1,padx=2,pady=2,sticky="nsew")
verify_data=Label(lf4,text="Verify & Update TimeTable",width=15)
verify_data.grid(row=1,column=2,columnspan=2, padx=2,pady=2,sticky="nsew")
verify_button=Button(lf4,text="Verified OK",bg="white",command=verify_ok)
verify_button.grid(row=2,column=2,padx=2,pady=2,sticky="nsew")
update_button=Button(lf4,text="Update TimeTable",bg="white",command=update_tt)
update_button.grid(row=2,column=3,padx=2,pady=2,sticky="nsew")
root.mainloop()